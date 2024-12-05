import cv2, os, numpy as np
import tkinter as tk
from PIL import ImageTk, Image
from datetime import datetime
from openpyxl import Workbook, load_workbook
from tkinter import PhotoImage

def selesai1():
    intructions.config(text="Rekam Data Telah Selesai!")
def selesai2():
    intructions.config(text="Training Wajah Telah Selesai!")
def selesai3():
    intructions.config(text="Absensi Telah Dilakukan")

def rekamDataWajah():
    wajahDir = 'datawajah'
    cam = cv2.VideoCapture(0)
    cam.set(3, 640)
    cam.set(4, 480)
    faceDetector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
    eyeDetector = cv2.CascadeClassifier('haarcascade_eye.xml')
    faceID = entry2.get()
    nama = entry1.get()
    nim = entry2.get()
    kelas = entry3.get()
    jurusan = entry4.get()
    ambilData = 1
    while True:
        retV, frame = cam.read()
        abuabu = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = faceDetector.detectMultiScale(abuabu, 1.3, 5)
        for (x, y, w, h) in faces:
            frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 255), 2)
            namaFile = str(nim) +'_'+str(nama) + '_' + str(kelas) +'_'+ str(jurusan) + '_' + str(ambilData) +'.jpg'
            cv2.imwrite(wajahDir + '/' + namaFile, frame)
            ambilData += 1
            roiabuabu = abuabu[y:y + h, x:x + w]
            roiwarna = frame[y:y + h, x:x + w]
            eyes = eyeDetector.detectMultiScale(roiabuabu)
            for (xe, ye, we, he) in eyes:
                cv2.rectangle(roiwarna, (xe, ye), (xe + we, ye + he), (0, 255, 255), 1)
        cv2.imshow('webcamku', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):  # jika menekan tombol q akan berhenti
            break
        elif ambilData > 30:
            break
    selesai1()
    cam.release()
    cv2.destroyAllWindows()  # untuk menghapus data yang sudah dibaca

def trainingWajah():
    wajahDir = 'datawajah'
    latihDir = 'latihwajah'

    def getImageLabel(path):
        imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
        faceSamples = []
        faceIDs = []
        for imagePath in imagePaths:
            PILimg = Image.open(imagePath).convert('L')
            imgNum = np.array(PILimg, 'uint8')
            faceID = int(os.path.split(imagePath)[-1].split('_')[0])
            faces = faceDetector.detectMultiScale(imgNum)
            for (x, y, w, h) in faces:
                faceSamples.append(imgNum[y:y + h, x:x + w])
                faceIDs.append(faceID)
            return faceSamples, faceIDs

    faceRecognizer = cv2.face.LBPHFaceRecognizer_create()
    faceDetector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
    faces, IDs = getImageLabel(wajahDir)
    faceRecognizer.train(faces, np.array(IDs))
    # simpan
    faceRecognizer.write(latihDir + '/training.xml')
    selesai2()

def markAttendance(name):
    filename = "Attendance.xlsx"
    try:
        # Jika file sudah ada, buka dan tambahkan data
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        # Jika file belum ada, buat baru dan tambahkan header
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Nama", "Kelas", "NIM","Jurusan", "Waktu"])

    yournim = entry2.get()
    yourclass = entry3.get()
    yourMajor = entry4.get()

    now = datetime.now()
    dtString = now.strftime('%Y-%m-%d %H:%M:%S')

    # Tambahkan data ke baris baru
    sheet.append([name, yourclass, yournim, yourMajor, dtString])

    # Simpan perubahan
    workbook.save(filename)
    workbook.close()

def absensiWajah():
    wajahDir = 'datawajah'
    latihDir = 'latihwajah'
    cam = cv2.VideoCapture(0)
    cam.set(3, 640)
    cam.set(4, 480)
    faceDetector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
    faceRecognizer = cv2.face.LBPHFaceRecognizer_create()
    faceRecognizer.read(latihDir + '/training.xml')
    font = cv2.FONT_HERSHEY_SIMPLEX

    #id = 0
    yourname = entry1.get()
    names = []
    names.append(yourname)
    minWidth = 0.1 * cam.get(3)
    minHeight = 0.1 * cam.get(4)

    while True:
        retV, frame = cam.read()
        frame = cv2.flip(frame, 1)
        abuabu = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = faceDetector.detectMultiScale(abuabu, 1.2, 5, minSize=(round(minWidth), round(minHeight)), )
        for (x, y, w, h) in faces:
            frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0),2)
            id, confidence = faceRecognizer.predict(abuabu[y:y+h,x:x+w])
            if (confidence < 100):
                id = names[0]
                confidence = "  {0}%".format(round(150 - confidence))
            elif confidence < 50:
                id = names[0]
                confidence = "  {0}%".format(round(170 - confidence))

            elif confidence > 70:
                id = "Tidak Diketahui"
                confidence = "  {0}%".format(round(150 - confidence))

            cv2.putText(frame, str(id), (x + 5, y - 5), font, 1, (255, 255, 255), 2)
            cv2.putText(frame, str(confidence), (x + 5, y + h + 25), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 0), 2)

        cv2.imshow('ABSENSI WAJAH', frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):  # jika menekan tombol q akan berhenti
            break
    markAttendance(id)
    selesai3()
    cam.release()
    cv2.destroyAllWindows()

# GUI
import tkinter as tk
from PIL import Image, ImageTk  # Importing Pillow for image manipulation

# Create the main window
root = tk.Tk()

# Create the canvas (window tkinter)
canvas = tk.Canvas(root, width=700, height=450)
canvas.grid(columnspan=3, rowspan=8)
canvas.configure(bg="#28a745")

# Load the logo using Pillow and resize it
image = Image.open("./img/UNPERBOS.png")  # Update this path to your logo's location
image = image.resize((70, 70))  # Resize the image to 50x50 pixels (adjust as needed)
logo = ImageTk.PhotoImage(image)  # Convert the resized image to a format tkinter can use

# Add the resized logo to the canvas at position (500, 500)
logo_label = tk.Label(root, image=logo, bg="#28a745")
canvas.create_window(90, 80, window=logo_label)  # Position it at the top-left corner of the canvas

# Title
judul = tk.Label(root, text="Smart Absensi", font=("Montserrat", 35, "bold"), bg="#28a745", fg="white")
canvas.create_window(350, 80, window=judul)

# Entry fields and labels for the form
entry1 = tk.Entry(root, font="Roboto")
canvas.create_window(457, 170, height=25, width=411, window=entry1)
label1 = tk.Label(root, text="Nama Siswa", font=("Roboto", 15, "bold"), fg="white", bg="#28a745")
canvas.create_window(110, 170, window=label1)

entry2 = tk.Entry(root, font="Roboto")
canvas.create_window(457, 210, height=25, width=411, window=entry2)
label2 = tk.Label(root, text="NIM", font=("Roboto", 15, "bold"), fg="white", bg="#28a745")
canvas.create_window(70, 210, window=label2)

entry3 = tk.Entry(root, font="Roboto")
canvas.create_window(457, 250, height=25, width=411, window=entry3)
label4 = tk.Label(root, text="Kelas", font=("Roboto", 15, "bold"), fg="white", bg="#28a745")
canvas.create_window(80, 250, window=label4)

entry4 = tk.Entry(root, font="Roboto")
canvas.create_window(457, 290, height=25, width=411, window=entry4)
label4 = tk.Label(root, text="Jurusan", font=("Roboto", 15, "bold"), fg="white", bg="#28a745")
canvas.create_window(90, 290, window=label4)

global intructions
# Instructions label
# tombol untuk rekam data wajah
intructions = tk.Label(root, text="Selamat Datang", font=("Roboto",15, "bold"),fg="white",bg="#28a745")
canvas.create_window(370, 350, window=intructions)
Rekam_text = tk.StringVar()
Rekam_btn = tk.Button(root, textvariable=Rekam_text, font="Roboto" , bg="white", fg="black", height=1, width=15,command=rekamDataWajah)
Rekam_text.set("Ambil Gambar")
Rekam_btn.grid(column=0, row=7)

# tombol untuk training wajah
Rekam_text1 = tk.StringVar()
Rekam_btn1 = tk.Button(root, textvariable=Rekam_text1, font="Roboto", bg="white", fg="black", height=1, width=15,command=trainingWajah)
Rekam_text1.set("Latih Wajah")
Rekam_btn1.grid(column=1, row=7)

# tombol absensi dengan wajah
Rekam_text2 = tk.StringVar()
Rekam_btn2 = tk.Button(root, textvariable=Rekam_text2, font="Roboto", bg="white", fg="black", height=1, width=20, command=absensiWajah)
Rekam_text2.set("Absen")
Rekam_btn2.grid(column=2, row=7)

# Run the tkinter loop
root.mainloop()